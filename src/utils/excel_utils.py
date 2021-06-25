from typing import *
from openpyxl import *
from openpyxl.styles import *
from openpyxl.utils import get_column_letter

def getCellStr(row, col):
    return f'{get_column_letter(col)}{row}'

def getRange(begin: Tuple[int, int], end: Tuple[int, int]):
    # given begin & end, return representation of merging argument
    # begin & end both in form of (row, col), note end is non-inclusive
    # e.g. begin=(1,2), end=(4,7) -> 'B1:F3'
    return f'{getCellStr(begin[0], begin[1])}:{getCellStr(end[0]-1, end[1]-1)}'

# print(getRange((1,2),(4,7)))

def makeBlock(
            ws, first: Tuple[int, int], title_sz: Tuple[int, int],
            col_widths: List[int], row_height: int, row_cnt: int
        ) -> Tuple[List[int], List[int]]:
        # all given in form (row, col), return list of column indexs
    # title cell
    center_aligned = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(getRange( (first[0], first[1]), (first[0]+title_sz[0], first[1]+title_sz[1]) ))
    factory_cell = ws[getCellStr(*first)]
    factory_cell.font = Font(size=18)
    factory_cell.alignment = center_aligned
    factory_cell.value = 'FactoryName'

    # merge cells
    # first column
    col_idx = []
    cur_col = first[1] + title_sz[1]
    for i, width in enumerate(col_widths):
        col_idx.append(cur_col)
        ws.merge_cells(getRange( (first[0], cur_col), (first[0]+title_sz[0], cur_col+width) ))
        current_cell = ws[getCellStr(first[0], cur_col)]
        current_cell.alignment = center_aligned
        current_cell.value = f'Hello{i}'
        cur_col += width

    # other column
    row_idx = []
    cur_row = first[0] + title_sz[0]
    for row_iter in range(row_cnt):
        row_idx.append(cur_row)
        cur_col = first[1] + title_sz[1] - 1
        for i, width in enumerate([1] + col_widths):
            ws.merge_cells(getRange( (cur_row, cur_col), (cur_row+row_height, cur_col+width) ))
            current_cell = ws[getCellStr(cur_row, cur_col)]
            current_cell.alignment = center_aligned
            if i == 0:
                current_cell.value = row_iter + 1
            cur_col += width
        cur_row += row_height
    
    return row_idx, col_idx
